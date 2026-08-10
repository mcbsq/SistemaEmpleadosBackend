# core/reportes.py
# ─────────────────────────────────────────────────────────────────────────────
# Generación de reportes — se pueden ver en línea (tabla en la propia app) o
# exportar a .xlsx. Cada `_datos_*` retorna (headers, rows, resumen):
#   - headers/rows: la tabla fila-por-fila, como antes.
#   - resumen: agregados que responden "quién / cómo / cuántos" de un vistazo,
#     sin tener que leer la tabla completa — pedido explícito del cliente tras
#     ver que los reportes solo listaban filas sin ningún análisis arriba.
# Tanto la vista en línea (JSON) como el .xlsx se construyen desde las mismas
# filas, así nunca se desincronizan.
from io import BytesIO
from datetime import date, datetime
from statistics import median

from openpyxl import Workbook
from openpyxl.styles import Font

from api.nomina.logic import calcular_nomina_dict
from api.vacaciones.logic import calcular_balance, _antiguedad_anios

# Catálogo de reportes — id estable (se usa en analitica_permisos), nombre y
# descripción para la UI. Agregar un reporte nuevo aquí lo hace disponible
# automáticamente para asignar permisos por rol.
CATALOGO_REPORTES = [
    {"id": "headcount", "nombre": "Headcount por departamento", "descripcion": "Quién compone cada área, con desglose por puesto y antigüedad promedio."},
    {"id": "nomina_resumen", "nombre": "Resumen de nómina", "descripcion": "Masa salarial total, promedio/mediana y quién gana más o menos.", "sensible": True},
    {"id": "vacaciones_uso", "nombre": "Uso de vacaciones", "descripcion": "Quién tiene días por perder, tasa de aprobación y próximas salidas."},
    {"id": "desempeno_resumen", "nombre": "Resumen de desempeño", "descripcion": "Brecha autoevaluación vs. jefe y % de evaluaciones completadas."},
    {"id": "reclutamiento", "nombre": "Reclutamiento", "descripcion": "Conversión por etapa del pipeline y antigüedad de las vacantes abiertas."},
]


def _autoajustar_columnas(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)


def _construir_wb(titulo, headers, rows, resumen=None):
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    fila = 1
    if resumen:
        for etiqueta, valor in resumen.items():
            ws.cell(row=fila, column=1, value=etiqueta).font = Font(bold=True)
            ws.cell(row=fila, column=2, value=valor)
            fila += 1
        fila += 1  # renglón en blanco antes de la tabla de detalle

    ws.append(headers) if fila == 1 else [ws.cell(row=fila, column=i + 1, value=h) for i, h in enumerate(headers)]
    for cell in ws[fila]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)

    _autoajustar_columnas(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _nombre_empleado(emp):
    return f"{emp.get('Nombre','')} {emp.get('ApelPaterno','')}".strip()


def _normalizar_area(valor):
    """'Sin Asignar' y 'Sin asignar' deben contar como la misma área — el
    cliente reportó esta duplicación por mayúsculas en el reporte real."""
    v = (valor or "Sin asignar").strip()
    return "Sin asignar" if v.lower() == "sin asignar" else v


# ─── 1. Headcount ───────────────────────────────────────────────────────────
def _datos_headcount(mongo):
    headers = ["Departamento", "Total empleados", "% del total", "Puestos", "Antigüedad promedio (años)"]
    empleados = list(mongo.db.empleados.find({"estado": {"$ne": "pendiente"}}))
    total = len(empleados)
    rh_por_emp = {str(r["empleado_id"]): r for r in mongo.db.rh.find()}

    por_area = {}
    for e in empleados:
        area = _normalizar_area(e.get("depto_id"))
        eid = str(e.get("_id"))
        rh = rh_por_emp.get(eid, {})
        grupo = por_area.setdefault(area, {"total": 0, "puestos": {}, "antiguedades": []})
        grupo["total"] += 1
        puesto = rh.get("Puesto")
        if puesto:
            grupo["puestos"][puesto] = grupo["puestos"].get(puesto, 0) + 1
        antig = _antiguedad_anios(rh.get("FechaIngreso"))
        if antig:
            grupo["antiguedades"].append(antig[0])

    rows = []
    for area, g in sorted(por_area.items(), key=lambda kv: kv[1]["total"], reverse=True):
        puestos_txt = ", ".join(f"{p} ({n})" for p, n in sorted(g["puestos"].items(), key=lambda kv: -kv[1])) or "Sin puestos registrados"
        antig_prom = round(sum(g["antiguedades"]) / len(g["antiguedades"]), 1) if g["antiguedades"] else None
        pct = round(g["total"] / total * 100) if total else 0
        rows.append([area, g["total"], f"{pct}%", puestos_txt, antig_prom if antig_prom is not None else "—"])

    resumen = {
        "Total de empleados": total,
        "Áreas registradas": len(por_area),
        "Área más grande": max(por_area.items(), key=lambda kv: kv[1]["total"])[0] if por_area else "—",
    }
    return headers, rows, resumen


# ─── 2. Nómina ──────────────────────────────────────────────────────────────
def _datos_nomina_resumen(mongo):
    headers = ["Empleado", "Puesto", "Percepción bruta", "ISR", "IMSS", "Neto mensual"]
    empleados = {str(e["_id"]): e for e in mongo.db.empleados.find()}
    rh_por_emp = {str(r["empleado_id"]): r for r in mongo.db.rh.find()}

    filas_calc = []
    for eid, rh in rh_por_emp.items():
        if (rh.get("TipoRelacionLaboral") or "nomina") == "prestador_servicios":
            continue
        calc = calcular_nomina_dict(mongo, eid)
        if not calc:
            continue
        emp = empleados.get(eid, {})
        filas_calc.append({
            "nombre": _nombre_empleado(emp) or eid,
            "puesto": rh.get("Puesto") or "Sin puesto",
            **calc,
        })

    rows = [[f["nombre"], f["puesto"], f["percepcion_bruta"], f["isr"], f["imss"], f["neto"]] for f in filas_calc]

    if filas_calc:
        netos = [f["neto"] for f in filas_calc]
        brutos = [f["percepcion_bruta"] for f in filas_calc]
        deducciones_total = sum(f["isr"] + f["imss"] for f in filas_calc)
        top = max(filas_calc, key=lambda f: f["neto"])
        bottom = min(filas_calc, key=lambda f: f["neto"])
        resumen = {
            "Masa salarial neta total": round(sum(netos), 2),
            "Promedio neto": round(sum(netos) / len(netos), 2),
            "Mediana neta": round(median(netos), 2),
            "ISR + IMSS como % del bruto": f"{round(deducciones_total / sum(brutos) * 100, 1)}%" if sum(brutos) else "0%",
            "Mayor percepción neta": f"{top['nombre']} (${top['neto']:,.2f})",
            "Menor percepción neta": f"{bottom['nombre']} (${bottom['neto']:,.2f})",
        }
    else:
        resumen = {"Empleados en nómina con salario configurado": 0}

    return headers, rows, resumen


# ─── 3. Vacaciones ──────────────────────────────────────────────────────────
def _datos_vacaciones_uso(mongo):
    headers = ["Empleado", "Solicitudes", "Días aprobados", "Días pendientes", "Días rechazados", "Días disponibles hoy"]
    empleados = list(mongo.db.empleados.find({"estado": {"$ne": "pendiente"}}))
    nombre_por_id = {str(e["_id"]): _nombre_empleado(e) for e in empleados}

    resumen_por_emp = {}
    for sol in mongo.db.vacaciones_solicitudes.find():
        eid = str(sol["empleado_id"])
        r = resumen_por_emp.setdefault(eid, {"solicitudes": 0, "aprobados": 0, "pendientes": 0, "rechazados": 0})
        r["solicitudes"] += 1
        dias = sol.get("dias_solicitados", 0)
        if sol.get("estado") == "aprobada":
            r["aprobados"] += dias
        elif sol.get("estado") == "pendiente":
            r["pendientes"] += dias
        else:
            r["rechazados"] += dias

    rows = []
    sin_solicitudes = []
    for e in empleados:
        eid = str(e["_id"])
        r = resumen_por_emp.get(eid, {"solicitudes": 0, "aprobados": 0, "pendientes": 0, "rechazados": 0})
        balance_resp = calcular_balance(mongo, eid)
        balance = balance_resp[0].get_json() if isinstance(balance_resp, tuple) else balance_resp.get_json()
        disponibles = balance.get("dias_disponibles", 0)
        rows.append([nombre_por_id[eid], r["solicitudes"], r["aprobados"], r["pendientes"], r["rechazados"], disponibles])
        if r["solicitudes"] == 0:
            sin_solicitudes.append(nombre_por_id[eid])

    rows.sort(key=lambda row: row[5], reverse=True)  # más días disponibles primero = mayor riesgo de perderlos

    total_solicitudes = sum(r["solicitudes"] for r in resumen_por_emp.values())
    total_aprobadas = sum(1 for sol in mongo.db.vacaciones_solicitudes.find() if sol.get("estado") == "aprobada")
    total_rechazadas = sum(1 for sol in mongo.db.vacaciones_solicitudes.find() if sol.get("estado") == "rechazada")
    evaluadas = total_aprobadas + total_rechazadas
    tasa_aprobacion = f"{round(total_aprobadas / evaluadas * 100)}%" if evaluadas else "—"

    resumen = {
        "Solicitudes totales del año": total_solicitudes,
        "Tasa de aprobación": tasa_aprobacion,
        "Empleados sin ninguna solicitud este año": len(sin_solicitudes),
        "Quiénes no han solicitado": ", ".join(sin_solicitudes[:5]) + (f" y {len(sin_solicitudes)-5} más" if len(sin_solicitudes) > 5 else "") if sin_solicitudes else "—",
    }
    return headers, rows, resumen


# ─── 4. Desempeño ───────────────────────────────────────────────────────────
def _datos_desempeno_resumen(mongo):
    headers = ["Ciclo", "Empleado", "Autoevaluación", "Evaluación de jefe", "Brecha (auto − jefe)"]
    ciclo = mongo.db.ciclos_evaluacion.find_one(sort=[("creado_en", -1)])
    if not ciclo:
        return headers, [], {"Ciclos de evaluación creados": 0}

    empleados = {str(e["_id"]): _nombre_empleado(e) for e in mongo.db.empleados.find()}
    evaluaciones = list(mongo.db.evaluaciones.find({"ciclo_id": ciclo["_id"]}))

    rows = []
    brechas = []
    completadas = 0
    for ev in evaluaciones:
        eid = str(ev["empleado_id"])
        auto = ev.get("autoevaluacion", {})
        jefe = ev.get("evaluacion_jefe", {})
        auto_ok, jefe_ok = auto.get("completada"), jefe.get("completada")
        if auto_ok and jefe_ok:
            completadas += 1
            brecha = round(auto["puntaje"] - jefe["puntaje"], 1)
            brechas.append((empleados.get(eid, eid), brecha))
        else:
            brecha = None
        rows.append([
            ciclo["nombre"], empleados.get(eid, eid),
            auto.get("puntaje") if auto_ok else "—",
            jefe.get("puntaje") if jefe_ok else "—",
            brecha if brecha is not None else "—",
        ])

    total = len(evaluaciones)
    pct_completadas = f"{round(completadas / total * 100)}%" if total else "0%"
    mayor_sobreestimacion = max(brechas, key=lambda t: t[1]) if brechas else None
    mayor_subestimacion = min(brechas, key=lambda t: t[1]) if brechas else None

    resumen = {
        "Ciclo": ciclo["nombre"],
        "Evaluaciones completas (ambas partes)": f"{completadas} de {total} ({pct_completadas})",
        "Mayor autoestimación (se califica arriba del jefe)": f"{mayor_sobreestimacion[0]} (+{mayor_sobreestimacion[1]})" if mayor_sobreestimacion and mayor_sobreestimacion[1] > 0 else "—",
        "Mayor subestimación (se califica abajo del jefe)": f"{mayor_subestimacion[0]} ({mayor_subestimacion[1]})" if mayor_subestimacion and mayor_subestimacion[1] < 0 else "—",
    }
    return headers, rows, resumen


# ─── 5. Reclutamiento (nuevo) ───────────────────────────────────────────────
def _datos_reclutamiento(mongo):
    headers = ["Vacante", "Área", "Estado", "Candidatos", "Días abierta"]
    vacantes = list(mongo.db.vacantes.find())
    candidatos = list(mongo.db.candidatos.find())

    por_vacante = {}
    for c in candidatos:
        vid = str(c.get("vacante_id"))
        por_vacante.setdefault(vid, []).append(c)

    hoy = date.today()
    rows = []
    for v in vacantes:
        vid = str(v["_id"])
        cands = por_vacante.get(vid, [])
        fecha_apertura = v.get("creado_en") or v.get("fecha_apertura")
        dias_abierta = "—"
        if fecha_apertura:
            try:
                f = datetime.strptime(str(fecha_apertura)[:10], "%Y-%m-%d").date()
                dias_abierta = (hoy - f).days
            except ValueError:
                pass
        rows.append([v.get("titulo") or v.get("nombre") or "Sin título", v.get("depto_id") or "Sin área",
                     v.get("estado", "abierta"), len(cands), dias_abierta])

    rows.sort(key=lambda r: r[4] if isinstance(r[4], int) else -1, reverse=True)

    por_etapa = {}
    for c in candidatos:
        etapa = c.get("etapa", "Sin etapa")
        por_etapa[etapa] = por_etapa.get(etapa, 0) + 1

    abiertas = [v for v in vacantes if v.get("estado", "abierta") == "abierta"]
    dias_abiertas_vals = [r[4] for r in rows if isinstance(r[4], int) and r[4] >= 0]

    resumen = {
        "Vacantes abiertas": len(abiertas),
        "Candidatos totales": len(candidatos),
        "Distribución por etapa": ", ".join(f"{k}: {v}" for k, v in sorted(por_etapa.items(), key=lambda kv: -kv[1])) or "—",
        "Antigüedad promedio de vacantes abiertas (días)": round(sum(dias_abiertas_vals) / len(dias_abiertas_vals)) if dias_abiertas_vals else "—",
    }
    return headers, rows, resumen


DATOS_GENERADORES = {
    "headcount": _datos_headcount,
    "nomina_resumen": _datos_nomina_resumen,
    "vacaciones_uso": _datos_vacaciones_uso,
    "desempeno_resumen": _datos_desempeno_resumen,
    "reclutamiento": _datos_reclutamiento,
}


def obtener_datos_reporte(mongo, reporte_id):
    """Retorna {"headers", "rows", "resumen"} para la vista en línea."""
    fn = DATOS_GENERADORES.get(reporte_id)
    if not fn:
        return None
    headers, rows, resumen = fn(mongo)
    return {"headers": headers, "rows": rows, "resumen": resumen}


def generar_reporte_xlsx(mongo, reporte_id):
    """Retorna un BytesIO con el .xlsx del reporte, o None si el id no existe."""
    fn = DATOS_GENERADORES.get(reporte_id)
    if not fn:
        return None
    headers, rows, resumen = fn(mongo)
    titulo = next((r["nombre"] for r in CATALOGO_REPORTES if r["id"] == reporte_id), reporte_id)
    return _construir_wb(titulo, headers, rows, resumen)


# Compatibilidad con el nombre viejo usado por api/analitica/routes.py.
GENERADORES = {rid: (lambda mongo, rid=rid: generar_reporte_xlsx(mongo, rid)) for rid in DATOS_GENERADORES}
